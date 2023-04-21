from pya3 import *
import openpyxl
import xlwings as xw
import pandas as pd
from datetime import datetime, time, timedelta,date
from time import sleep
import threading
import schedule

class Login:

    def __init__(self, api, user_Id):
        self.alice = Aliceblue(api_key=api
                               , user_id=user_Id)
        print(self.alice.get_session_id())
        """ for getting previous day data use 9:00 if running at another time comment
         the first schedule from class D_1H - on_run()"""
        self.previous_day_data = datetime.strptime("9:00", "%H:%M").time()

        """For changing the time periods change the first four and do the same for rest four"""

        self.start_time_1 = datetime.strptime("10:45", "%H:%M").time()
        self.start_time_2 = datetime.strptime("11:00", "%H:%M").time()
        self.start_time_3 = datetime.strptime("11:15", "%H:%M").time()
        self.start_time_4 = datetime.strptime("11:30", "%H:%M").time()

        self.t1 = time(10, 45)
        self.t2 = time(11, 00)
        self.t3 = time(11, 15)
        self.t4 = time(11, 30)


class D_1H:

    def __init__(self, login_obj: Login):
        self.login_obj = login_obj
        self.sym = {
            "9:30" : [],
            "9:45" : [],
            "10:00" : [],
        }
        self.data_placer_1 = False
        self.data_placer_2 = False
        self.data_placer_3 = False
        self.lp_data = {}
        self.LTP = 0
        self.socket_opened = False
        self.subscribe_flag = False
        self.subscribe_list = []
        self.unsubscribe_list = []

    def _1ms(self):

        def socket_open():  # Socket open callback function
            print("Connected")
            # global socket_opened
            self.socket_opened = True
            if self.subscribe_flag:  # This is used to resubscribe the script when reconnect the socket.
                self.login_obj.alice.subscribe(self.subscribe_list)

        def socket_close():  # On Socket close this callback function will trigger
            # global socket_opened, LTP, lp_data
            self.socket_opened = False
            self.LTP = 0
            print("Closed")

        def socket_error(message):  # Socket Error Message will receive in this callback function
            # global LTP
            self.LTP = 0
            print("Error :", message)

        def feed_data(message):  # Socket feed data will receive in this callback function
            # global LTP, subscribe_flag, s, token_value
            feed_message = json.loads(message)
            lp_data_copy = self.lp_data.copy()

            if feed_message["t"] == "ck":
                print("Connection Acknowledgement status :%s (Websocket Connected)" % feed_message["s"])
                self.subscribe_flag = True
                print("subscribe_flag :", self.subscribe_flag)
                print("-------------------------------------------------------------------------------")
                pass
            elif feed_message["t"] == "tk":
                print("Token Acknowledgement status :%s " % feed_message)
                print("-------------------------------------------------------------------------------")
                pass
            else:
                # print("Feed :", feed_message)
                for token in self.token_value:
                    if feed_message.get('tk') == str(token):
                        self.lp_data[feed_message.get('tk')] = feed_message.get('lp')
                try:
                    for i in self.token_value:
                        value = self.lp_data[str(i)]
                        if value is None:
                            self.lp_data[str(i)] = lp_data_copy[str(i)]

                except:
                    pass
                # print(self.lp_data)

        # Socket Connection Request
        self.login_obj.alice.start_websocket(socket_open_callback=socket_open, socket_close_callback=socket_close,
                              socket_error_callback=socket_error, subscription_callback=feed_data,
                              run_in_background=True, market_depth=False)

        while not self.socket_opened:
            pass
        workbook = openpyxl.load_workbook('TradingExcel.xlsx')
        worksheet = workbook['LiveData']
        symbols = []
        for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
            for cell in row:
                symbols.append(cell.value)
        sym = list(filter(lambda x: x is not None, symbols))
        self.token_value = []
        for s in sym:
            self.subscribe_list = [self.login_obj.alice.get_instrument_by_symbol('NSE', s)]
            # print(self.subscribe_list)
            self.token_value.append(self.subscribe_list[0].token)

            self.login_obj.alice.subscribe(self.subscribe_list)
            sleep(3)
        # print(self.token_value)

        ##Program ko close nahi karna hai loop chalte renehe dena hai
        while True:
            pass
    def ltp_to_excel(self):
        while True:

            wb = xw.Book('TradingExcel.xlsx')
            ws = wb.sheets['LiveData']
            ltp = []
            # print(self.lp_data)
            for token, value in self.lp_data.items():
                ltp.append(value)
            ws.range("E4").options(transpose=True).value = ltp
            # wb.save()
            sleep(0.1)

    def run_soc(self):
        thread1 = threading.Thread(target=self._1ms)
        thread2 = threading.Thread(target=self.ltp_to_excel)
        thread1.start()
        thread2.start()
    def __1d(self):
        alice = self.login_obj.alice
        try:

            workbook = openpyxl.load_workbook('TradingExcel.xlsx')
            worksheet = workbook['LiveData']
            symbols = []
            for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
                for cell in row:
                    symbols.append(cell.value)
            sym = list(filter(lambda x: x is not None, symbols))

            wb = xw.Book("TradingExcel.xlsx")
            ws = wb.sheets['LiveData']
            stk_data = pd.DataFrame()

            # print(sym)
            self.sym["9:30"] = sym
            print(self.sym["9:30"])

            for symbol in sym:
                # print(symbol)
                if (datetime.now().time() >= time(9, 30)) and (datetime.now().time() <= time(15,10)):
                    instrument = alice.get_instrument_by_symbol("NSE", symbol)
                    # print(instrument)
                    from_datetime = datetime.now() - timedelta(days=1)
                    to_datetime = datetime.now()
                    interval = "1"  # ["1", "D"]
                    indices = False
                    df = alice.get_historical(instrument, from_datetime, to_datetime, interval, indices)
                    # print(df)
                    df.drop(["volume"], axis=1, inplace=True)
                    df['datetime'] = pd.to_datetime(df['datetime'])
                    df.set_index("datetime", inplace=True)
                    df.index = pd.to_datetime(df.index)
                    # print(df)
                    df_15min = df.resample('15T').agg({'open': 'first', 'high': 'max', 'low': 'min', 'close': 'last'})
                    df_15min.drop(["open", "close"], axis=1, inplace=True)
                    df_15min = df_15min.iloc[-1]
                    # print(df_15min)
                    # df_15min = df_15min.transpose()
                    # print(df_15min)
                    print("CheckPoint-1")
                    stk_data = pd.concat([stk_data, df_15min], axis=1)
                else:
                    instrument = alice.get_instrument_by_symbol("NSE", symbol)
                    from_datetime =datetime.now() - timedelta(days=2)
                    to_datetime = datetime.now()
                    interval = "D"  # ["1", "D"]
                    indices = False
                    df = alice.get_historical(instrument, from_datetime, to_datetime, interval, indices)
                    # print(type(df))
                    df.drop(["volume"], axis=1, inplace=True)
                    df['datetime'] = pd.to_datetime(df['datetime'])
                    df.set_index("datetime", inplace=True)
                    df.index = pd.to_datetime(df.index)
                    # df_15min = df.resample('15T').agg({'open': 'first', 'high': 'max', 'low': 'min', 'close': 'last'})
                    df = df.iloc[-1]
                    print(df)

                    # stk_data = pd.concat([stk_data, df_15min], axis=1)
                    stk_data = pd.concat([stk_data, df], axis=1)
            now = datetime.now().strftime("%H:%M")
            # print(now)
            target_time = self.login_obj.start_time_1
            if datetime.strptime(now, "%H:%M").time() == target_time:
                stk_data = stk_data.transpose()
                print(stk_data)
                ws.range("J3").options(index=False).value = stk_data
                wb.save()
                self.data_placer_1 = True
            target_time = self.login_obj.start_time_2
            if datetime.strptime(now, "%H:%M").time() == target_time:
                stk_data = stk_data.transpose()
                print(stk_data)
                ws.range("L3").options(index=False).value = stk_data
                wb.save()
                self.data_placer_2 = True
            target_time = self.login_obj.start_time_3
            if datetime.strptime(now, "%H:%M").time() == target_time:
                stk_data = stk_data.transpose()
                print(stk_data)
                ws.range("N3").options(index=False).value = stk_data
                wb.save()
                self.data_placer_3 = True
            now = datetime.now().time()
            if now <= time(9, 30):
                stk_data = stk_data.transpose()
                print(stk_data)
                ws.range("F3").options(index=False).value = stk_data
                wb.save()
        except Exception as e:
            print("Error in Hist_data: ",e)

    def on_run(self):
        # schedule.every().day.at(str(self.login_obj.previous_day_data)).do(self.__1d)
        schedule.every().day.at(str(self.login_obj.start_time_1)).do(self.__1d)
        schedule.every().day.at(str(self.login_obj.start_time_2)).do(self.__1d)
        schedule.every().day.at(str(self.login_obj.start_time_3)).do(self.__1d)

        # Run the scheduler continuously
        while True:
            schedule.run_pending()
            # print("Checkpoint: waiting for 15min data")
            sleep(1)

    def _1ds(self):
        thread = threading.Thread(target=self.on_run)
        thread.start()


class O_1P:

    def __init__(self, login_obj: Login, data_obj: D_1H):
        self.login_obj = login_obj
        self.data_obj = data_obj
        self.order_books = {
            "9:30":[],
            "9:45":[],
            "10:00":[],
        }

        self.order_ids = {
            "9:30":[],
            "9:45":[],
            "10:00":[]
                          }
        self.O1 = False
        self.O2 = False
        self.O3 = False
        self.orderplaced = 0
    def _1of(self):
        workbook = openpyxl.load_workbook('TradingExcel.xlsx')
        worksheet = workbook['LiveData']
        order_list = []
        symbols = []
        quantity = []
        filterd_sym = []
        filtered_high = []
        filtered_low = []
        for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
            for cell in row:
                symbols.append(cell.value)
        sym = list(filter(lambda x: x is not None, symbols))
        for row in worksheet.iter_rows(min_row=4,max_row=(4 + len(sym)),  min_col=3, max_col=3):
            for cell in row:
                quantity.append(cell.value)
        for row in worksheet.iter_rows(min_row=4, min_col=4, max_col=4):
            for cell in row:
                order_list.append(cell.value)
        checks = list(filter(lambda x: x is not None, order_list))
        highs = []
        lows = []
        qty = []
        now = datetime.now().strftime("%H:%M")
        target_time = self.login_obj.start_time_1
        if datetime.strptime(now, "%H:%M").time() == target_time and self.O1 == False and self.data_obj.data_placer_1 == True:
            for row in worksheet.iter_rows(min_row=4, min_col=10, max_col=10):
                for cell in row:
                    highs.append(cell.value)
            highs_List = list(filter(lambda x: x is not None, highs))
            for row in worksheet.iter_rows(min_row=4, min_col=11, max_col=11):
                for cell in row:
                    lows.append(cell.value)
            lows_List = list(filter(lambda x: x is not None, lows))
            for i in range(len(checks)):
                if checks[i] == "Y":
                    qty.append(quantity[i])
                    filterd_sym.append(sym[i])
                    filtered_high.append(highs_List[i])
                    filtered_low.append(lows_List[i])

            self.place(filtered_high, filterd_sym, filtered_low, qty)
            self.O1 = True
        now = datetime.now().strftime("%H:%M")
        target_time = self.login_obj.start_time_2
        if datetime.strptime(now, "%H:%M").time() == target_time and self.O2 == False and self.data_obj.data_placer_2 == True:
            print(self.data_obj.sym['9:45'])
            for row in worksheet.iter_rows(min_row=4, min_col=12, max_col=12):
                for cell in row:
                    highs.append(cell.value)
            highs_List = list(filter(lambda x: x is not None, highs))
            for row in worksheet.iter_rows(min_row=4, min_col=13, max_col=13):
                for cell in row:
                    lows.append(cell.value)
            lows_List = list(filter(lambda x: x is not None, lows))

            for i in range(len(sym)):
                if checks[i] == "Y" and sym[i] in self.data_obj.sym["9:45"]:
                    filterd_sym.append(sym[i])
                    qty.append(quantity[i])
                    filtered_high.append(highs_List[i])
                    filtered_low.append(lows_List[i])
            self.place(filtered_high, filterd_sym, filtered_low, qty)
            self.O2 = True
        now = datetime.now().strftime("%H:%M")
        target_time = self.login_obj.start_time_3
        if datetime.strptime(now, "%H:%M").time() == target_time and self.O3 == False and self.data_obj.data_placer_3 == True:
            print(self.data_obj.sym['10:00'])
            for row in worksheet.iter_rows(min_row=4, min_col=14, max_col=14):
                for cell in row:
                    highs.append(cell.value)
            highs_List = list(filter(lambda x: x is not None, highs))
            for row in worksheet.iter_rows(min_row=4, min_col=15, max_col=15):
                for cell in row:
                    lows.append(cell.value)
            lows_List = list(filter(lambda x: x is not None, lows))
            for i in range(len(sym)):
                if checks[i] == "Y" and sym[i] in self.data_obj.sym["10:00"]:
                        filterd_sym.append(sym[i])
                        qty.append(quantity[i])
                        filtered_high.append(highs_List[i])
                        filtered_low.append(lows_List[i])
            print(filterd_sym, filtered_high, filtered_low)
            self.place(filtered_high, filterd_sym, filtered_low, qty)
            self.O3 = True
    def place(self, high_list, sym, low_list, qty):
        alice = self.login_obj.alice
        for symbol, high, low, quantity in zip(sym, high_list, low_list, qty):
            order_1 = alice.place_order(transaction_type = TransactionType.Buy,
                                 instrument = alice.get_instrument_by_symbol('NSE', symbol),
                                 quantity = quantity,
                                 order_type = OrderType.StopLossLimit,
                                 product_type = ProductType.Intraday,
                                 price = float(high),
                                 trigger_price = float(high),
                                 stop_loss = 0,
                                 square_off = None,
                                 trailing_sl = None,
                                 is_amo = False)
            print(order_1)
            order_2 = alice.place_order(transaction_type=TransactionType.Sell,
                              instrument=alice.get_instrument_by_symbol('NSE', symbol),
                              quantity=quantity,
                              order_type=OrderType.StopLossLimit,
                              product_type=ProductType.Intraday,
                              price=float(low),
                              trigger_price=float(low),
                              stop_loss=0,
                              square_off=None,
                              trailing_sl=None,
                              is_amo=False)
            print(order_2)

            now = datetime.now().time()
            if now >= self.login_obj.t1 and now < self.login_obj.t2:
                self.order_ids["9:30"].append(order_1["NOrdNo"])
                self.order_ids["9:30"].append(order_2["NOrdNo"])
            if now >= self.login_obj.t2 and now < self.login_obj.t3:
                self.order_ids["9:45"].append(order_1['NOrdNo'])
                self.order_ids["9:45"].append(order_2['NOrdNo'])
            if now >= self.login_obj.t3 and now < self.login_obj.t4:
                self.order_ids["10:00"].append(order_1['NOrdNo'])
                self.order_ids["10:00"].append(order_2['NOrdNo'])

    def _1os(self):
        alice = self.login_obj.alice
        df = pd.DataFrame()
        wb = xw.Book("TradingExcel.xlsx")
        ws = wb.sheets['Orders']
        list_of_dicts = []
        now = datetime.now().strftime("%H:%M")
        start_time = self.login_obj.start_time_1
        end_time = self.login_obj.start_time_2
        if start_time <= datetime.strptime(now, "%H:%M").time() < end_time:
            if self.order_ids["9:30"]:
                for id in self.order_ids["9:30"]:
                    list_of_dicts.append(alice.get_order_history(id))
                for d in list_of_dicts:
                    row = {
                        'Exchange': d['Exchange'],
                        'Symbol': d['Sym'],
                        'Time': d['OrderedTime'],
                        'Qty': d['Qty'],
                        "Type": d['Trantype'],
                        'OrderType': d['Prctype'],
                        'LimitPrice': d['Prc'],
                        'TriggerPrice': d['Trgprc'],
                        'Status': d['Status'],
                        'OrderNo': d['Nstordno']
                    }
                    df = pd.concat([df, pd.DataFrame(row, index=[0])], ignore_index=True)

                self.order_books['9:30'].append(df)
                ws.range("A3").options(index=False, header=False).value = df
                wb.save()
        now = datetime.now().strftime("%H:%M")
        start_time = self.login_obj.start_time_2
        end_time = self.login_obj.start_time_3
        if start_time <= datetime.strptime(now, "%H:%M").time() < end_time:
            if self.order_ids["9:45"]:
                for id in self.order_ids["9:45"]:
                    list_of_dicts.append(alice.get_order_history(id))
                for d in list_of_dicts:
                    row = {
                        'Exchange': d['Exchange'],
                        'Symbol': d['Sym'],
                        'Time': d['OrderedTime'],
                        'Qty': d['Qty'],
                        "Type": d['Trantype'],
                        'OrderType': d['Prctype'],
                        'LimitPrice': d['Prc'],
                        'TriggerPrice': d['Trgprc'],
                        'Status': d['Status'],
                        'OrderNo': d['Nstordno']
                    }
                    df = pd.concat([df, pd.DataFrame(row, index=[0])], ignore_index=True)
                self.order_books['9:45'].append(df)
                ws.range("AC3").options(index=False, header=False).value = df
                wb.save()
        now = datetime.now().strftime("%H:%M")
        start_time = self.login_obj.start_time_3
        end_time = self.login_obj.start_time_4
        if start_time <= datetime.strptime(now, "%H:%M").time() < end_time:
            if self.order_ids["10:00"]:
                for id in self.order_ids["10:00"]:
                    list_of_dicts.append(alice.get_order_history(id))
                for d in list_of_dicts:
                    row = {
                        'Exchange': d['Exchange'],
                        'Symbol': d['Sym'],
                        'Time': d['OrderedTime'],
                        'Qty': d['Qty'],
                        "Type": d['Trantype'],
                        'OrderType': d['Prctype'],
                        'LimitPrice': d['Prc'],
                        'TriggerPrice': d['Trgprc'],
                        'Status': d['Status'],
                        'OrderNo': d['Nstordno']
                    }
                    df = pd.concat([df, pd.DataFrame(row, index=[0])], ignore_index=True)
                self.order_books['10:00'].append(df)
                ws.range("BE3").options(index=False, header=False).value = df
                wb.save()

class T_2S:
    def __init__(self, login_obj: Login):
        self.login_obj = login_obj
        self.order_id_sl = {
            "9:30": [],
            "9:45": [],
            "10:00": [],
        }
        self.cmp_sym = {
            "9:30": [],
            "9:45": [],
            "10:00": [],
        }

    def sl_target_order(self, sym):

        SL = None
        Target = None

        now = datetime.now().time()
        if now >= self.login_obj.t1 and now <= self.login_obj.t2:

            df = pd.read_excel("TradingExcel.xlsx", sheet_name="Orders", header=None)

            subset_df = df[[10, 11]]
            SL_rows = subset_df[subset_df[10] == '(SL)']
            if not SL_rows.empty:
                SL_index = SL_rows.index[0]
                SL = subset_df.iloc[SL_index + 1:][subset_df.columns[0]].tolist()
            Target_rows = subset_df[subset_df[11] == 'Target']
            if not Target_rows.empty:
                Target_index = Target_rows.index[0]
                Target = subset_df.iloc[Target_index + 1:][subset_df.columns[1]].tolist()
        now = datetime.now().time()
        if now >= self.login_obj.t2 and now <= self.login_obj.t3:

            df = pd.read_excel("TradingExcel.xlsx", sheet_name="Orders", header=None)
            subset_df = df[[38, 39]]
            SL_rows = subset_df[subset_df[38] == '(SL)']
            if not SL_rows.empty:
                SL_index = SL_rows.index[0]
                SL = subset_df.iloc[SL_index + 1:][subset_df.columns[0]].tolist()

            Target_rows = subset_df[subset_df[39] == 'Target']
            if not Target_rows.empty:
                Target_index = Target_rows.index[0]
                Target = subset_df.iloc[Target_index + 1:][subset_df.columns[1]].tolist()
        now = datetime.now().time()
        if now >= self.login_obj.t3 and now <= self.login_obj.t4:

            df = pd.read_excel("TradingExcel.xlsx", sheet_name="Orders", header=None)
            subset_df = df[[66, 67]]
            SL_rows = subset_df[subset_df[66] == '(SL)']
            if not SL_rows.empty:
                SL_index = SL_rows.index[0]
                SL = subset_df.iloc[SL_index + 1:][subset_df.columns[0]].tolist()

            Target_rows = subset_df[subset_df[67] == 'Target']
            if not Target_rows.empty:
                Target_index = Target_rows.index[0]
                Target = subset_df.iloc[Target_index + 1:][subset_df.columns[1]].tolist()
        self.place_slTg(sym, SL, Target)

    def place_slTg(self, ticker, sl, target):
        alice = self.login_obj.alice
        status = []
        symbols = []
        workbook = openpyxl.load_workbook('TradingExcel.xlsx')
        worksheet = workbook['Orders']
        sl_1 = None
        tg_1 = None
        side = None
        qty = None
        order_status = []
        quantity = []
        sym = []
        type = []
        trns_type = []
        now = datetime.now().time()
        if now >= self.login_obj.t1 and now <= self.login_obj.t2:
            if ticker not in self.cmp_sym["9:30"]:
                for row in worksheet.iter_rows(min_row=3, min_col=5, max_col=5):
                    for cell in row:
                        type.append(cell.value)
                trns_type = list(filter(lambda x: x is not None, type))
                for row in worksheet.iter_rows(min_row=3, min_col=9, max_col=9):
                    for cell in row:
                        status.append(cell.value)
                order_status = list(filter(lambda x: x is not None, status))
                for row in worksheet.iter_rows(min_row=3, min_col=2, max_col=2):
                    for cell in row:
                        symbols.append(cell.value)
                sym = list(filter(lambda x: x is not None, symbols))
                for row in worksheet.iter_rows(min_row=3, min_col=4, max_col=4):
                    for cell in row:
                        quantity.append(cell.value)
                self.cmp_sym["9:30"].append(ticker)
        if now >= self.login_obj.t2 and now <= self.login_obj.t3:

            if ticker not in self.cmp_sym["9:45"]:
                for row in worksheet.iter_rows(min_row=3, min_col=33, max_col=33):
                    for cell in row:
                        type.append(cell.value)
                trns_type = list(filter(lambda x: x is not None, type))
                for row in worksheet.iter_rows(min_row=3, min_col=37, max_col=37):
                    for cell in row:
                        status.append(cell.value)
                order_status = list(filter(lambda x: x is not None, status))
                for row in worksheet.iter_rows(min_row=3, min_col=30, max_col=30):
                    for cell in row:
                        symbols.append(cell.value)
                sym = list(filter(lambda x: x is not None, symbols))
                for row in worksheet.iter_rows(min_row=3, min_col=32, max_col=32):
                    for cell in row:
                        quantity.append(cell.value)
                self.cmp_sym["9:45"].append(ticker)

        if now >= self.login_obj.t3 and now <= self.login_obj.t4:
            if ticker not in self.cmp_sym["10:00"]:
                for row in worksheet.iter_rows(min_row=3, min_col=61, max_col=61):
                    for cell in row:
                        type.append(cell.value)
                trns_type = list(filter(lambda x: x is not None, type))
                for row in worksheet.iter_rows(min_row=3, min_col=65, max_col=65):
                    for cell in row:
                        status.append(cell.value)
                order_status = list(filter(lambda x: x is not None, status))
                for row in worksheet.iter_rows(min_row=3, min_col=58, max_col=58):
                    for cell in row:
                        symbols.append(cell.value)
                sym = list(filter(lambda x: x is not None, symbols))
                for row in worksheet.iter_rows(min_row=3, min_col=60, max_col=60):
                    for cell in row:
                        quantity.append(cell.value)
                self.cmp_sym["10:00"].append(ticker)
        # print(quantity)
        try:
            for i in range(len(order_status)):
                if order_status[i] == "complete":
                    complete_sym = sym[i]
                    side = trns_type[i]
                    qty = quantity[i]
                    print(qty)
                    if complete_sym == ticker:
                        roun_sl = sl[i]
                        sl_1 = round(roun_sl * 2) / 2
                        print(sl_1)

                        roun_tg = target[i]
                        tg_1 = round(roun_tg * 2) / 2

                        print(tg_1)

            if side == "B":
                order_1 = alice.place_order(transaction_type=TransactionType.Sell,
                                            instrument=alice.get_instrument_by_symbol('NSE', ticker),
                                            quantity=int(qty),
                                            order_type=OrderType.StopLossLimit,
                                            product_type=ProductType.Intraday,
                                            price=float(tg_1),
                                            trigger_price=float(tg_1),
                                            stop_loss=None,
                                            square_off=None,
                                            trailing_sl=None,
                                            is_amo=False,
                                            order_tag='order1')

                order_2 = alice.place_order(transaction_type=TransactionType.Sell,
                                            instrument=alice.get_instrument_by_symbol('NSE', ticker),
                                            quantity=int(qty),
                                            order_type=OrderType.StopLossLimit,
                                            product_type=ProductType.Intraday,
                                            price=float(sl_1),
                                            trigger_price=float(sl_1),
                                            stop_loss=None,
                                            square_off=None,
                                            trailing_sl=None,
                                            is_amo=False,
                                            order_tag='order1')

            else:
                order_1 = alice.place_order(transaction_type=TransactionType.Buy,
                                            instrument=alice.get_instrument_by_symbol('NSE', ticker),
                                            quantity=int(qty),
                                            order_type=OrderType.StopLossLimit,
                                            product_type=ProductType.Intraday,
                                            price=float(tg_1),
                                            trigger_price=float(tg_1),
                                            stop_loss=None,
                                            square_off=None,
                                            trailing_sl=None,
                                            is_amo=False,
                                            order_tag='order1')

                order_2 = alice.place_order(transaction_type=TransactionType.Buy,
                                            instrument=alice.get_instrument_by_symbol('NSE', ticker),
                                            quantity=int(qty),
                                            order_type=OrderType.StopLossLimit,
                                            product_type=ProductType.Intraday,
                                            price=float(sl_1),
                                            trigger_price=float(sl_1),
                                            stop_loss=None,
                                            square_off=None,
                                            trailing_sl=None,
                                            is_amo=False,
                                            order_tag='order1')
            print(order_1)
            print(order_2)

            now = datetime.now().time()
            if now >= self.login_obj.t1 and now <= self.login_obj.t2:
                self.order_id_sl["9:30"].append(order_1['NOrdNo'])
                self.order_id_sl["9:30"].append(order_2['NOrdNo'])
                print(self.order_id_sl["9:30"])

            if now >= self.login_obj.t2 and now <= self.login_obj.t3:
                self.order_id_sl["9:45"].append(order_1['NOrdNo'])
                self.order_id_sl["9:45"].append(order_2['NOrdNo'])
                print(self.order_id_sl["9:45"])
            if now >= self.login_obj.t3 and now <= self.login_obj.t4:
                self.order_id_sl["10:00"].append(order_1['NOrdNo'])
                self.order_id_sl["10:00"].append(order_2['NOrdNo'])
                print(self.order_id_sl["10:00"])
        except:
            pass

    def _1tt(self):

        df_1 = pd.DataFrame()
        df_2 = pd.DataFrame()
        df_3 = pd.DataFrame()
        wb = xw.Book("TradingExcel.xlsx")
        ws = wb.sheets['Orders']
        list_of_dicts = []

        if self.order_id_sl["9:30"]:
            for id in self.order_id_sl["9:30"]:
                list_of_dicts.append(self.login_obj.alice.get_order_history(id))
            for d in list_of_dicts:
                row = {
                    'Exchange': d['Exchange'],
                    'Symbol': d['Sym'],
                    'Time': d['OrderedTime'],
                    'Qty': d['Qty'],
                    "Type": d['Trantype'],
                    'OrderType': d['Prctype'],
                    'LimitPrice': d['Prc'],
                    'TriggerPrice': d['Trgprc'],
                    'Status': d['Status'],
                    'OrderNo': d['Nstordno']
                }
                df_1 = pd.concat([df_1, pd.DataFrame(row, index=[0])], ignore_index=True)

        if self.order_id_sl["9:45"]:
            for id in self.order_id_sl["9:45"]:
                list_of_dicts.append(self.login_obj.alice.get_order_history(id))
            for d in list_of_dicts:
                row = {
                    'Exchange': d['Exchange'],
                    'Symbol': d['Sym'],
                    'Time': d['OrderedTime'],
                    'Qty': d['Qty'],
                    "Type": d['Trantype'],
                    'OrderType': d['Prctype'],
                    'LimitPrice': d['Prc'],
                    'TriggerPrice': d['Trgprc'],
                    'Status': d['Status'],
                    'OrderNo': d['Nstordno']
                }
                df_2 = pd.concat([df_2, pd.DataFrame(row, index=[0])], ignore_index=True)

        if self.order_id_sl["10:00"]:
            for id in self.order_id_sl["10:00"]:
                list_of_dicts.append(self.login_obj.alice.get_order_history(id))
            for d in list_of_dicts:
                row = {
                    'Exchange': d['Exchange'],
                    'Symbol': d['Sym'],
                    'Time': d['OrderedTime'],
                    'Qty': d['Qty'],
                    "Type": d['Trantype'],
                    'OrderType': d['Prctype'],
                    'LimitPrice': d['Prc'],
                    'TriggerPrice': d['Trgprc'],
                    'Status': d['Status'],
                    'OrderNo': d['Nstordno']
                }
                df_3 = pd.concat([df_3, pd.DataFrame(row, index=[0])], ignore_index=True)

        now = datetime.now().time()
        if now >= self.login_obj.t1:
            ws.range("O3").options(index=False, header=False).value = df_1
            wb.save()
        if now >= self.login_obj.t2:
            ws.range("AQ3").options(index=False, header=False).value = df_2
            wb.save()
        if now >= self.login_obj.t3:
            ws.range("BS3").options(index=False, header=False).value = df_3
            wb.save()

class T_3O:
    def __init__(self, login_obj: Login, data_obj: D_1H, sl_tg: T_2S):
        self.login_obj = login_obj
        self.data_obj = data_obj
        self.slTg = sl_tg


    def __4sc(self):
        alice = self.login_obj.alice
        status = []
        symbols = []
        order_id = []
        workbook = openpyxl.load_workbook('TradingExcel.xlsx')
        worksheet = workbook['Orders']
        try:
            now = datetime.now().strftime("%H:%M")
            start_time = self.login_obj.start_time_1
            end_time = self.login_obj.start_time_2
            if start_time <= datetime.strptime(now, "%H:%M").time() <= end_time:
                for row in worksheet.iter_rows(min_row=3, min_col=9, max_col=9):
                    for cell in row:
                        status.append(cell.value)
                order_status = list(filter(lambda x: x is not None, status))
                for row in worksheet.iter_rows(min_row=3, min_col=2, max_col=2):
                    for cell in row:
                        symbols.append(cell.value)
                sym = list(filter(lambda x: x is not None, symbols))

                for row in worksheet.iter_rows(min_row=3, min_col=10, max_col=10):
                    for cell in row:
                        order_id.append(cell.value)
                id = list(filter(lambda x: x is not None, order_id))
                cancle_id = None
                if order_status:
                    for i in range(len(order_status)):
                        if order_status[i] == "complete":

                           complete_sym = sym[i]
                           complete_id = id[i]
                           if complete_sym in self.data_obj.sym["9:30"]:
                               self.slTg.sl_target_order(complete_sym)
                           else:
                               pass

                           for j in range(len(sym)):
                               if sym[j] == complete_sym and id[j] != complete_id:
                                   cancle_id = id[j]
                                   alice.cancel_order(cancle_id)
                               else:
                                   pass
                        start = self.login_obj.start_time_2
                        new_time = datetime.combine(datetime.today(), start) - timedelta(minutes=1)
                        result = new_time.time()
                        now = datetime.now().time().strftime("%H:%M")
                        if datetime.strptime(now, "%H:%M").time() >= result:
                            if order_status[i] == "trigger pending" or order_status[i] == "open":
                                self.data_obj.sym["9:45"].append(sym[i])
                                self.data_obj.sym["9:45"] = list(dict.fromkeys(self.data_obj.sym["9:45"]))
                                cancle_id = id[i]
                                alice.cancel_order(cancle_id)
                else:
                    pass
        except Exception as e:
            print("Error in tracker1 9:30-45: ", e)

        try:

            now = datetime.now().time().strftime("%H:%M")
            start_time = self.login_obj.start_time_2
            end_time = self.login_obj.start_time_3
            if start_time <= datetime.strptime(now, "%H:%M").time() <= end_time:
                for row in worksheet.iter_rows(min_row=3, min_col=37, max_col=37):
                    for cell in row:
                        status.append(cell.value)
                order_status = list(filter(lambda x: x is not None, status))
                for row in worksheet.iter_rows(min_row=3, min_col=30, max_col=30):
                    for cell in row:
                        symbols.append(cell.value)
                sym = list(filter(lambda x: x is not None, symbols))

                for row in worksheet.iter_rows(min_row=3, min_col=38, max_col=38):
                    for cell in row:
                        order_id.append(cell.value)
                id = list(filter(lambda x: x is not None, order_id))
                cancle_id = None
                if order_status:
                    for i in range(len(order_status)):
                        if order_status[i] == "complete":

                            complete_sym = sym[i]
                            complete_id = id[i]
                            if complete_sym in self.data_obj.sym["9:45"]:
                                self.slTg.sl_target_order(complete_sym)
                            else:
                                pass

                            for j in range(len(sym)):
                                if sym[j] == complete_sym and id[j] != complete_id:
                                    cancle_id = id[j]
                                    alice.cancel_order(cancle_id)
                                else:
                                    pass
                        start = self.login_obj.start_time_3
                        new_time = datetime.combine(datetime.today(), start) - timedelta(minutes=1)
                        result = new_time.time()
                        now = datetime.now().time().strftime("%H:%M")
                        if datetime.strptime(now, "%H:%M").time() == result:
                            if order_status[i] == "trigger pending" or order_status[i] == "open":
                                self.data_obj.sym["10:00"].append(sym[i])
                                self.data_obj.sym["10:00"] = list(dict.fromkeys(self.data_obj.sym["10:00"]))
                                cancle_id = id[i]
                                alice.cancel_order(cancle_id)
                else:
                    pass

        except Exception as e:
            print("Error in Tarcker2 9:45-60: ", e)

        try:
            now = datetime.now().time().strftime("%H:%M")
            start_time = self.login_obj.start_time_3
            end_time = self.login_obj.start_time_4
            if start_time <= datetime.strptime(now, "%H:%M").time() <= end_time:
                # print("In condition")
                for row in worksheet.iter_rows(min_row=3, min_col=65, max_col=65):
                    for cell in row:
                        status.append(cell.value)
                order_status = list(filter(lambda x: x is not None, status))
                # print(order_status)
                for row in worksheet.iter_rows(min_row=3, min_col=58, max_col=58):
                    for cell in row:
                        symbols.append(cell.value)
                sym = list(filter(lambda x: x is not None, symbols))
                # print(sym)
                for row in worksheet.iter_rows(min_row=3, min_col=66, max_col=66):
                    for cell in row:
                        order_id.append(cell.value)
                id = list(filter(lambda x: x is not None, order_id))
                cancle_id = None
                if order_status:
                    for i in range(len(order_status)):
                        if order_status[i] == "complete":

                            complete_sym = sym[i]
                            complete_id = id[i]
                            if complete_sym in self.data_obj.sym["10:00"]:
                                self.slTg.sl_target_order(complete_sym)
                            else:
                                pass

                            for j in range(len(sym)):
                                if sym[j] == complete_sym and id[j] != complete_id:
                                    cancle_id = id[j]
                                    alice.cancel_order(cancle_id)
                                else:
                                    pass
                        start = self.login_obj.start_time_4
                        new_time = datetime.combine(datetime.today(), start) - timedelta(minutes=1)
                        result = new_time.time()
                        now = datetime.now().time().strftime("%H:%M")
                        if datetime.strptime(now, "%H:%M").time() == result:
                            if order_status[i] == "trigger pending" or order_status[i] == "open":
                                cancle_id = id[i]
                                alice.cancel_order(cancle_id)
                else:
                    pass
        except Exception as e:
            print("Error in tracker3 10-15: ", e)
        status = []
        symbols = []
        order_id = []
        now = datetime.now().time()
        if now >= self.login_obj.t1:
            for row in worksheet.iter_rows(min_row=3, min_col=23, max_col=23):
                for cell in row:
                    status.append(cell.value)
            order_status = list(filter(lambda x: x is not None, status))
            for row in worksheet.iter_rows(min_row=3, min_col=16, max_col=16):
                for cell in row:
                    symbols.append(cell.value)
            sym = list(filter(lambda x: x is not None, symbols))
            for row in worksheet.iter_rows(min_row=3, min_col=24, max_col=24):
                for cell in row:
                    order_id.append(cell.value)
            id = list(filter(lambda x: x is not None, order_id))
            cancle_id = None

            for i in range(len(order_status)):
                if order_status[i] == "complete":
                    complete_sym = sym[i]
                    complete_id = id[i]

                    for j in range(len(sym)):
                        if sym[j] == complete_sym and id[j] != complete_id:
                            cancle_id = id[j]
                            alice.cancel_order(cancle_id)
                        else:
                            pass
        now = datetime.now().time()
        if now >= self.login_obj.t2:
            for row in worksheet.iter_rows(min_row=3, min_col=51, max_col=51):
                for cell in row:
                    status.append(cell.value)
            order_status = list(filter(lambda x: x is not None, status))
            for row in worksheet.iter_rows(min_row=3, min_col=44, max_col=44):
                for cell in row:
                    symbols.append(cell.value)
            sym = list(filter(lambda x: x is not None, symbols))
            for row in worksheet.iter_rows(min_row=3, min_col=52, max_col=52):
                for cell in row:
                    order_id.append(cell.value)
            id = list(filter(lambda x: x is not None, order_id))
            cancle_id = None

            for i in range(len(order_status)):
                if order_status[i] == "complete":
                    complete_sym = sym[i]
                    complete_id = id[i]

                    for j in range(len(sym)):
                        if sym[j] == complete_sym and id[j] != complete_id:
                            cancle_id = id[j]
                            alice.cancel_order(cancle_id)
                        else:
                            pass
        now = datetime.now().time()
        if now >= self.login_obj.t3:
            for row in worksheet.iter_rows(min_row=3, min_col=79, max_col=79):
                for cell in row:
                    status.append(cell.value)
            order_status = list(filter(lambda x: x is not None, status))
            for row in worksheet.iter_rows(min_row=3, min_col=72, max_col=72):
                for cell in row:
                    symbols.append(cell.value)
            sym = list(filter(lambda x: x is not None, symbols))
            for row in worksheet.iter_rows(min_row=3, min_col=80, max_col=80):
                for cell in row:
                    order_id.append(cell.value)
            id = list(filter(lambda x: x is not None, order_id))
            cancle_id = None

            for i in range(len(order_status)):
                if order_status[i] == "complete":
                    complete_sym = sym[i]
                    complete_id = id[i]
                    for j in range(len(sym)):
                        if sym[j] == complete_sym and id[j] != complete_id:
                            cancle_id = id[j]
                            alice.cancel_order(cancle_id)
                            print(cancle_id)
                        else:
                            pass

    def __5ts(self):
        alice = self.login_obj.alice
        workbook = openpyxl.load_workbook('TradingExcel.xlsx')
        worksheet = workbook['Orders']
        status = []
        order = []
        """time 9:30"""
        now = datetime.now().strftime("%H:%M")
        start_time = self.login_obj.start_time_1
        end_time =self.login_obj.start_time_2
        if start_time <= datetime.strptime(now, "%H:%M").time() <= end_time:
            for row in worksheet.iter_rows(min_row=3, min_col=13, max_col=13):
                for cell in row:
                    status.append(cell.value)
            action_upper= list(filter(lambda x: x is not None, status))
            action = [elem.lower() for elem in action_upper]
            for row in worksheet.iter_rows(min_row=3, min_col=10, max_col=10):
                for cell in row:
                    order.append(cell.value)
            OrderId = list(filter(lambda x: x is not None, order))
            # print(OrderId)
            order_no = []
            if action:
                for i in range(len(action)):
                    if action[i] == "exit":
                        order_no.append(OrderId[i])
                if order_no:
                    for no in order_no:
                        try:
                            alice.cancel_order(no)
                            # print("order cancle")
                        except:
                            pass
            else:
                pass
        now = datetime.now().strftime("%H:%M")
        start_time = self.login_obj.start_time_2
        end_time = self.login_obj.start_time_3
        if start_time <= datetime.strptime(now, "%H:%M").time() <= end_time:
            for row in worksheet.iter_rows(min_row=3, min_col=41, max_col=41):
                for cell in row:
                    status.append(cell.value)
            action_upper = list(filter(lambda x: x is not None, status))
            action = [elem.lower() for elem in action_upper]
            for row in worksheet.iter_rows(min_row=3, min_col=38, max_col=38):
                for cell in row:
                    order.append(cell.value)
            OrderId = list(filter(lambda x: x is not None, order))
            # print(OrderId)
            order_no = []
            if action:
                for i in range(len(action)):
                    if action[i] == "exit":
                        order_no.append(OrderId[i])
                if order_no:
                    for no in order_no:
                        try:
                            alice.cancel_order(no)
                            # print("order cancle")
                        except:
                            pass
            else:
                pass
        now = datetime.now().strftime("%H:%M")
        start_time = self.login_obj.start_time_3
        end_time = self.login_obj.start_time_4
        if start_time <= datetime.strptime(now, "%H:%M").time() <= end_time:
            for row in worksheet.iter_rows(min_row=3, min_col=69, max_col=69):
                for cell in row:
                    status.append(cell.value)
            action_upper = list(filter(lambda x: x is not None, status))
            action = [elem.lower() for elem in action_upper]
            for row in worksheet.iter_rows(min_row=3, min_col=66, max_col=66):
                for cell in row:
                    order.append(cell.value)
            OrderId = list(filter(lambda x: x is not None, order))
            # print(OrderId)
            if now == time(10, 45):
                for i in range(len(OrderId)):
                    alice.cancel_order(i)
                print("All orders from order list of 10:00 are cancle")

            else:
                order_no = []
                if action:
                    for i in range(len(action)):
                        if action[i] == "exit":
                            order_no.append(OrderId[i])

                    if order_no:
                        for no in order_no:
                            try:
                                alice.cancel_order(no)
                                # print("order cancle")
                            except:
                                pass
                else:
                    pass
        status = []
        order = []

        now = datetime.now().time()
        if now >= self.login_obj.t1:
            for row in worksheet.iter_rows(min_row=3, min_col=27, max_col=27):
                for cell in row:
                    status.append(cell.value)
            action_upper = list(filter(lambda x: x is not None, status))
            action = [elem.lower() for elem in action_upper]
            for row in worksheet.iter_rows(min_row=3, min_col=24, max_col=24):
                for cell in row:
                    order.append(cell.value)
            OrderId = list(filter(lambda x: x is not None, order))
            # print(OrderId)
            order_n = []
            if action:
                for i in range(len(action)):
                    if action[i] == "exit":
                        order_n.append(OrderId[i])
                if order_n:
                    for no in order_n:
                        try:
                            alice.cancel_order(no)
                            # print("order cancle")
                        except:
                            pass
            else:
                pass
        now = datetime.now().time()
        if now >= self.login_obj.t2:
            for row in worksheet.iter_rows(min_row=3, min_col=55, max_col=55):
                for cell in row:
                    status.append(cell.value)
            action_upper = list(filter(lambda x: x is not None, status))
            action = [elem.lower() for elem in action_upper]
            for row in worksheet.iter_rows(min_row=3, min_col=52, max_col=52):
                for cell in row:
                    order.append(cell.value)
            OrderId = list(filter(lambda x: x is not None, order))
            # print(OrderId)
            order_n = []
            if action:
                for i in range(len(action)):
                    if action[i] == "exit":
                        order_n.append(OrderId[i])
                if order_n:
                    for no in order_n:
                        try:
                            alice.cancel_order(no)
                        except:
                            pass
            else:
                pass
        now = datetime.now().time()
        if now >= self.login_obj.t3:
            for row in worksheet.iter_rows(min_row=3, min_col=83, max_col=83):
                for cell in row:
                    status.append(cell.value)
            action_upper = list(filter(lambda x: x is not None, status))
            action = [elem.lower() for elem in action_upper]
            for row in worksheet.iter_rows(min_row=3, min_col=80, max_col=80):
                for cell in row:
                    order.append(cell.value)
            OrderId = list(filter(lambda x: x is not None, order))
            # print(OrderId)
            order_n = []
            if action:
                for i in range(len(action)):
                    if action[i] == "exit":
                        order_n.append(OrderId[i])
                if order_n:
                    for no in order_n:
                        try:
                            alice.cancel_order(no)
                        except:
                            pass
            else:
                pass
    def run(self):
        thread1 = threading.Thread(target=self.__4sc)
        thread2 = threading.Thread(target=self.__5ts)
        thread1.start()
        thread2.start()

class Time:
    def __init__(self):
        # xw.App()
        self.wb = xw.Book("TradingExcel.xlsx")
        self.ws = self.wb.sheets['LiveData']
    def T_E(self):
        now = datetime.now().time()
        now_str = now.strftime('%H:%M:%S')
        self.ws.range('B1').value = now_str
        self.ws.range('B1').number_format = 'hh:mm:ss'



time_obj = Time()

login_details = open("login_details.txt", "r")
aa =login_details.readlines()

api = aa[0].strip('\n')
user_ID = aa[1].strip('\n')

login_obj = Login(api, user_ID)

data_obj = D_1H(login_obj)

data_obj._1ds()
data_obj.run_soc()
order_obj = O_1P(login_obj,data_obj)
sl_tg = T_2S(login_obj)
order_track = T_3O(login_obj, data_obj,sl_tg)

while True:
    if datetime.now().time() >= login_obj.t1 and datetime.now().time() <= time(15,30):
        try:
            time_obj.T_E()
        except Exception as e:
                print("Error in time to excel",e)
        try:
            order_obj._1of()
        except Exception as e:
            print("Error in Order fetcher",e)

        order_obj._1os()
        try:
            order_track.run()
        except Exception as e:
            print("Error in Order tracker",e)
        try:
            sl_tg._1tt()
        except Exception as e:
            print("Error in Sl target",e)

        sleep(1)

